from django.conf import settings
from openpyxl.styles.colors import BLUE

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, Color

from apps.map.models import Region
from apps.ratings.models import MONTHS, MonthlyRating, MonthlyRatingElement

center_align = Alignment(horizontal='center', vertical='center')
halign_left_valign_center_wrap = Alignment(horizontal='left',
                                           vertical='center',
                                           wrap_text=True)
halign_center_valign_center_wrap = Alignment(horizontal='center',
                                             vertical='center',
                                             wrap_text=True)
center_align_wrap = Alignment(horizontal='center', vertical='center',
                              wrap_text=True)
align_wrap = Alignment(wrap_text=True)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
medium_border = Border(left=Side(style='medium'),
                       right=Side(style='medium'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))

main_header_font = Font(name='Calibri',
                        size=16,
                        bold=True)

table_header_font = Font(name='Calibri',
                         size=12,
                         bold=True)

hyperlink_font = Font(name='Calibri',
                      size=12,
                      color=Color(rgb=BLUE))

number_format_1 = '0.0'
number_format_2 = '0.00'
number_format_percent = '0.00%'

light_blue_fill = PatternFill("solid", fgColor='{}{}{}'.format(hex(221)[2:], hex(235)[2:], hex(246)[2:]))
blue_fill = PatternFill("solid", fgColor='{}{}{}'.format(hex(94)[2:], hex(156)[2:], hex(211)[2:]))
light_gray_fill = PatternFill("solid", fgColor='{}{}{}'.format(hex(217)[2:], hex(217)[2:], hex(217)[2:]))
light_green_fill = PatternFill("solid", fgColor='{}{}{}'.format(hex(199)[2:], hex(223)[2:], hex(182)[2:]))

hex_0 = format(0, '02x')
hex_60 = format(60, '02x')
hex_120 = format(120, '02x')
hex_220 = format(220, '02x')
hex_255 = format(255, '02x')

red_fill = PatternFill("solid", fgColor='{}{}{}'.format(hex(255)[2:], hex_0, hex_0))
green_fill = PatternFill("solid", fgColor='{}{}{}'.format(hex_0, hex(255)[2:], hex_0))


def get_color_fill(value: float, min_val: float, max_val: float, type: int) -> PatternFill:
    if value is not None:
        result = None
        if min_val == max_val:
            result = PatternFill("solid", fgColor='{}{}{}'.format(hex_0, hex_220, hex_120))
        else:
            new_max = max_val - min_val
            new_val = value - min_val
            if new_max != 0:
                multiplexer = 1 / abs(new_max)
            else:
                multiplexer = 1
            if multiplexer != 1:
                new_val *= multiplexer
            # print(value, new_max, min_val, max_val)
            if type == 1:
                if new_val <= 0.5:
                    result = PatternFill(
                        "solid",
                        fgColor='{}{}{}'.format(
                            format(int(new_val * 2 * 255), '02x'),
                            hex_220,
                            hex_0)
                    )
                else:
                    result = PatternFill(
                        "solid",
                        fgColor='{}{}{}'.format(
                            hex_220,
                            format(int((1 - new_val) * 2 * 255), '02x'),
                            hex_0)
                    )
            elif type == 2:
                if new_val <= 0.5:
                    result = PatternFill(
                        "solid",
                        fgColor='{}{}{}'.format(
                            hex_220,
                            format(int(new_val * 2 * 255), '02x'),
                            hex_0)
                    )
                else:
                    result = PatternFill(
                        "solid",
                        fgColor='{}{}{}'.format(
                            format(int((1 - new_val) * 2 * 255), '02x'),
                            hex_220,
                            hex_0)
                    )
        return result


class MonthlyRatingExcelGenerator:
    def __init__(self, monthly_rating: MonthlyRating):
        self.monthly_rating = monthly_rating
        self.rating_elements = monthly_rating.elements.all()\
            .prefetch_related('rating_element')\
            .prefetch_related('related_sub_elements')\
            .prefetch_related('rating_element__base_document') \
            .prefetch_related('related_sub_elements__values')
        self.rating_elements_values = {element.id: element.values for element in self.rating_elements}
        self.rating_elements_weighted_values = {}
        for element in self.rating_elements:
            weighted_values = {
                value: self.rating_elements_values[element.id][value] * element.rating_element.weight
                       if self.rating_elements_values[element.id][value] is not None
                       else None
                for value in element.values
            }
            self.rating_elements_weighted_values[element.id] = weighted_values
        self.regions = list(Region.objects.all())
        self.wb = Workbook()
        self.initial_ws = self.wb.get_active_sheet()

    def generate(self) -> Workbook:
        self.generate_sheets()
        self.wb.remove(self.initial_ws)
        return self.wb

    def generate_sheets(self) -> None:
        rating_ws = self.wb.create_sheet(MONTHS[self.monthly_rating.month].lower())
        self.fill_main_sheet(rating_ws)
        for element in self.rating_elements:
            if element.related_sub_elements.exists():
                element_ws = self.wb.create_sheet()
                element_ws.title = '{}'.format(element.number)
                self.fill_element_sheet(element, element_ws)

    def get_sum_values(self) -> dict:
        sum_values = {region.id: None for region in self.regions}
        for region in self.regions:
            for element_value in self.rating_elements_weighted_values:
                if self.rating_elements_weighted_values[element_value][region.id] is not None:
                    if sum_values[region.id] is not None:
                        sum_values[region.id] += float(self.rating_elements_weighted_values[element_value][region.id])
                    else:
                        sum_values[region.id] = float(self.rating_elements_weighted_values[element_value][region.id])
        return sum_values

    def get_regions_places(self, sum_values: dict) -> dict:
        return {k: idx + 1
                for idx, (k, v)
                in enumerate(reversed(sorted([v for v in sum_values.items() if v[1] is not None], key=lambda x: x[1])))
                if v is not None}

    def get_max_possible_value(self) -> int:
        max_sum = 0
        for element in self.rating_elements:
            max_sum += element.rating_element.weight
        return max_sum

    ###########################################################################
    # MAIN SHEET
    ###########################################################################

    def fill_main_sheet(self, sheet) -> None:
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                    'O', 'P', 'Q', 'R']:
            sheet.column_dimensions[col].width = 7

        sheet.column_dimensions['S'].width = 10
        sheet.column_dimensions['T'].width = 35
        sheet.column_dimensions['U'].width = 30
        sheet.column_dimensions['V'].width = 30

        row_offset = 1
        # Main header
        cell = sheet.cell(row=row_offset, column=20)
        cell.value = self.monthly_rating.signer_text.text
        cell.font = table_header_font
        cell.alignment = center_align_wrap
        row_offset += 1

        cell = sheet.cell(row=row_offset, column=1)
        cell.value = \
            'Исходные данные для расчета рейтинга управ районов САО ' \
            'в части показателей ЖКХ за {month} {year} года в соответствии с ' \
            '{document}'.format(
                month=MONTHS[self.monthly_rating.month].lower(),
                year=self.monthly_rating.year,
                document=self.monthly_rating.base_document.description_imp
            )
        cell.font = table_header_font
        cell.alignment = center_align_wrap
        cell.fill = light_blue_fill
        sheet.merge_cells('A2:T2')
        row_offset += 2

        sum_values = self.get_sum_values()
        regions_places = self.get_regions_places(sum_values)
        column_offset = 2
        cell = sheet.cell(row=row_offset, column=column_offset)
        cell.value = 'место'
        cell.alignment = center_align
        cell.font = table_header_font
        column_offset += 1

        for idx, region in enumerate(self.regions):
            cell = sheet.cell(row=row_offset, column=column_offset + idx)
            cell.value = regions_places[region.id] if region.id in regions_places else None
            cell.alignment = center_align
            cell.font = table_header_font
        row_offset += 1

        # Table headers
        table_headers_cells_1 = [sheet.cell(row=row_offset, column=idx + 1)
                                 for idx in range(22)]
        table_headers_values_1 = [
            'Наименование показателей',
            'Ответственный',
        ]
        for region in self.regions:
            table_headers_values_1.append(region.short_name)
        table_headers_values_1.append('Средний')
        table_headers_values_1.append('Описание показателя')
        table_headers_values_1.append('Комментарии согласовывающего')
        table_headers_values_1.append('Информация о\nзамечаниях районов')

        for idx, cell in enumerate(table_headers_cells_1):
            cell.value = table_headers_values_1[idx]
            cell.alignment = center_align_wrap
            cell.border = medium_border
            cell.font = table_header_font
        row_offset += 1

        table_headers_cells_2 = [sheet.cell(row=row_offset, column=idx + 1)
                                 for idx in range(22)]
        table_headers_cells_2[0].value = 'Суммарный показатель'
        for idx, cell in enumerate(table_headers_cells_2):
            cell.alignment = center_align
            cell.border = medium_border
            cell.font = table_header_font
            cell.number_format = number_format_1

        if sum_values:
            min_sum_value = min([v for k, v in sum_values.items() if v is not None])
            max_sum_value = max([v for k, v in sum_values.items() if v is not None])
            for idx, region in enumerate(self.regions):
                if sum_values[region.id] is not None:
                    table_headers_cells_2[idx + 2].value = sum_values[region.id]
                    table_headers_cells_2[idx + 2].fill = get_color_fill(sum_values[region.id], min_sum_value, max_sum_value, 2)
                    table_headers_cells_2[18].value = sum([val[1] for val in sum_values.items() if val[1] is not None]) / len(sum_values)
        row_offset += 1

        table_headers_cells_3 = [sheet.cell(row=row_offset, column=idx + 1)
                                 for idx in range(22)]
        table_headers_cells_3[0].value = 'Максимально возможный\nсуммарный показатель'
        max_possible_value = self.get_max_possible_value()
        for idx in range(len(self.regions)):
            table_headers_cells_3[idx + 2].value = max_possible_value
        for idx, cell in enumerate(table_headers_cells_3):
            cell.alignment = center_align_wrap
            cell.border = medium_border
            cell.font = table_header_font
            cell.number_format = number_format_1
        row_offset += 1

        for idx, element in enumerate(self.rating_elements):
            column_offset = 1
            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value='{}) {}'.format(idx + 1, element.rating_element.name)
            )
            cell.border = thin_border
            cell.alignment = halign_left_valign_center_wrap
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=element.responsible.short_name if element.responsible else ''
            )
            cell.border = thin_border
            cell.alignment = center_align
            column_offset += 1

            vals_list = [float(v)
                         for k, v in self.rating_elements_values[element.id].items()
                         if v is not None]

            if vals_list:
                min_sum_value = float(min(vals_list)) * element.rating_element.weight
                max_sum_value = float(max(vals_list)) * element.rating_element.weight

            for region in self.regions:
                cell = sheet.cell(
                    row=idx + row_offset,
                    column=column_offset,
                )
                val = self.rating_elements_weighted_values[element.id][region.id]
                cell.value = val
                cell.border = thin_border
                cell.alignment = center_align
                cell.number_format = number_format_1
                if val is not None:
                    cell.fill = get_color_fill(float(val), min_sum_value, max_sum_value, 2)
                column_offset += 1

            cell_val = None
            if vals_list:
                cell_val = (sum(vals_list)/len(vals_list)) * element.rating_element.weight
            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=cell_val
            )
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = number_format_1
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
            )
            val = element.rating_element.base_description if element.rating_element.base_description else ''
            val += element.additional_description if element.additional_description else ''
            cell.value = val
            cell.border = thin_border
            cell.alignment = halign_left_valign_center_wrap
            cell.fill = light_blue_fill
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=element.negotiator_comment
            )
            cell.border = thin_border
            cell.alignment = halign_left_valign_center_wrap
            cell.fill = red_fill
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=element.region_comment
            )
            cell.border = thin_border
            cell.alignment = center_align_wrap

    ###########################################################################
    # ELEMENT SHEET
    ###########################################################################

    def fill_element_sheet(self, element: MonthlyRatingElement, sheet):
        sheet.column_dimensions['A'].width = 45
        sheet.column_dimensions['B'].width = 20
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                    'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']:
            sheet.column_dimensions[col].width = 8

        sheet.column_dimensions['W'].width = 45
        sheet.column_dimensions['X'].width = 25

        row_offset = 1

        # Main header
        sheet.merge_cells('A1:X1')
        sheet.row_dimensions[1].height = 45
        cell = sheet.cell(row=row_offset, column=1)
        cell.value = \
            'Исходные данные для расчета комплексного показателя №{number} ' \
            '"{name}"\nрейтинга управ районов САО в части показателей ЖКХ ' \
            'за {month} {year} года в соответствии с {document}'.format(
                number=element.number,
                name=element.rating_element.name,
                month=MONTHS[self.monthly_rating.month].lower(),
                year=self.monthly_rating.year,
                document=element.rating_element.base_document.description_imp
            )
        cell.font = table_header_font
        cell.alignment = center_align_wrap
        cell.fill = light_blue_fill
        row_offset += 2

        table_headers_cells_1 = [sheet.cell(row=row_offset, column=idx + 1)
                                 for idx in range(24)]
        table_headers_values_1 = [
            'Наименование\nпоказателей',
            'Ответственный',
        ]
        for region in self.regions:
            table_headers_values_1.append(region.short_name)
        table_headers_values_1.append('лучш')
        table_headers_values_1.append('лучш')
        table_headers_values_1.append('мин')
        table_headers_values_1.append('макс')
        table_headers_values_1.append('Описание показателя')
        table_headers_values_1.append('ссылка на документ\nоснование')

        for idx, cell in enumerate(table_headers_cells_1):
            cell.value = table_headers_values_1[idx]
            cell.alignment = center_align_wrap
            cell.border = medium_border
            cell.fill = blue_fill
            cell.font = table_header_font
        row_offset += 1

        table_headers_cells_2 = [sheet.cell(row=row_offset, column=idx + 1)
                                 for idx in range(24)]
        table_headers_cells_2[0].value = 'Итоговый комплексный показатель'
        table_headers_cells_2[1].value = element.responsible.short_name

        min_val = None
        max_val = None
        for idx, region in enumerate(self.regions):
            val = self.rating_elements_values[element.id][region.id]
            table_headers_cells_2[idx + 2].value = val
            if val is not None and min_val is not None and val < min_val:
                min_val = val
            elif min_val is None and val is not None:
                min_val = val
            if val is not None and max_val is not None and val > max_val:
                max_val = val
            elif max_val is None and val is not None:
                max_val = val
            table_headers_cells_2[idx + 2].number_format = number_format_2
        for idx, region in enumerate(self.regions):
            if table_headers_cells_2[idx + 2].value is not None:
                table_headers_cells_2[idx + 2].fill = get_color_fill(
                    table_headers_cells_2[idx + 2].value,
                    min_val, max_val,
                    2
                )
        table_headers_cells_2[20].value = min_val
        table_headers_cells_2[20].fill = red_fill
        table_headers_cells_2[21].value = max_val
        table_headers_cells_2[21].fill = green_fill

        for idx, cell in enumerate(table_headers_cells_2):
            cell.alignment = center_align
            cell.border = thin_border
            cell.number_format = number_format_2
        row_offset += 1

        table_headers_cells_3 = [sheet.cell(row=row_offset, column=idx + 1)
                                 for idx in range(24)]
        for idx, cell in enumerate(table_headers_cells_3):
            cell.value = idx + 1
            cell.fill = light_gray_fill
            cell.font = table_header_font
            cell.alignment = center_align
        row_offset += 1

        sub_elements = element.related_sub_elements.all()
        for idx, sub_element in enumerate(sub_elements):
            list_values = \
                [list(value)
                 for value
                 in sub_element.values.values_list('region', 'is_average', 'value')]

            min_val = sub_element.get_min_value(list_values, 2)
            max_val = sub_element.get_max_value(list_values, 2)
            avg_val = sub_element.get_avg_value(list_values, 2)

            if sub_element.display_type == 1:
                number_format = number_format_2
            elif sub_element.display_type == 2:
                number_format = number_format_percent

            column_offset = 1
            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=sub_element.name
            )
            cell.border = thin_border
            cell.alignment = halign_left_valign_center_wrap
            cell.fill = light_blue_fill
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=sub_element.responsible.short_name
            )
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = light_blue_fill
            column_offset += 1

            for region in self.regions:
                cell = sheet.cell(
                    row=idx + row_offset,
                    column=column_offset,
                )
                value = sub_element.values.get(region=region.id)
                if value.is_average:
                    cell.value = avg_val
                    cell.fill = light_gray_fill
                else:
                    cell.value = value.value
                    if value.value is not None:
                        cell.fill = get_color_fill(value.value, min_val, max_val, sub_element.best_type)
                cell.number_format = number_format
                cell.border = thin_border
                cell.alignment = center_align

                column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=sub_element.get_best_type_display()
            )
            cell.border = thin_border
            cell.alignment = center_align
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=min_val
                      if sub_element.display_type == 1
                      else max_val
            )
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = number_format
            cell.fill = light_green_fill
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=min_val
            )
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = number_format
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=max_val
            )
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = number_format
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
                value=sub_element.description
            )
            cell.border = thin_border
            cell.alignment = halign_left_valign_center_wrap
            column_offset += 1

            cell = sheet.cell(
                row=idx + row_offset,
                column=column_offset,
            )
            cell.border = thin_border
            cell.alignment = halign_left_valign_center_wrap
            cell.font = hyperlink_font
            if sub_element.document:
                cell.value = '=HYPERLINK("{}/{}", "{}")'.format(
                    settings.BASE_URL,
                    sub_element.document,
                    sub_element.document.name[sub_element.document.name.rfind('/') + 1:]
                )
            column_offset += 1
